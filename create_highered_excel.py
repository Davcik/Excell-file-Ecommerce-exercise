
import pandas as pd
from openpyxl import Workbook

# Create a workbook and remove the default sheet
wb = Workbook()
ws = wb.active
wb.remove(ws)

# --- Cash Flow sheet ---
cash_flow = wb.create_sheet("Cash Flow")
cash_flow_data = [
    ["Month", "Revenue", "Expenses", "Net Cash Flow", "Cumulative Cash Flow"],
    ["Jan", "=SUM(ProductMix!B2:B10*Pricing!B2:B10)", 15000, "=B2-C2", "=D2"],
    ["Feb", "=SUM(ProductMix!C2:C10*Pricing!B2:B10)", 16000, "=B3-C3", "=E2+D3"],
    ["Mar", "=SUM(ProductMix!D2:D10*Pricing!B2:B10)", 17000, "=B4-C4", "=E3+D4"]
]
for row in cash_flow_data:
    cash_flow.append(row)

# --- Product Mix sheet ---
product_mix = wb.create_sheet("ProductMix")
product_mix_data = [
    ["Course", "Jan Sales", "Feb Sales", "Mar Sales"],
    ["Course A", 100, 120, 130],
    ["Course B", 80, 90, 95],
    ["Course C", 60, 70, 75],
    ["Course D", 40, 50, 55],
    ["Course E", 30, 35, 40],
    ["Course F", 20, 25, 30],
    ["Course G", 15, 20, 25],
    ["Course H", 10, 12, 15],
    ["Course I", 8, 10, 12]
]
for row in product_mix_data:
    product_mix.append(row)

# --- Pricing sheet ---
pricing = wb.create_sheet("Pricing")
pricing_data = [
    ["Course", "Price per Unit"],
    ["Course A", 500],
    ["Course B", 450],
    ["Course C", 400],
    ["Course D", 350],
    ["Course E", 300],
    ["Course F", 250],
    ["Course G", 200],
    ["Course H", 150],
    ["Course I", 100]
]
for row in pricing_data:
    pricing.append(row)

# Save the file
file_path = "HigherEd_Courses_Business_Model.xlsx"
wb.save(file_path)
print(f"Excel file saved as {file_path}")
